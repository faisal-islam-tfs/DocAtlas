import gzip
import io
import json
import logging
import subprocess
import tempfile
import unittest
import zipfile
from argparse import Namespace
from pathlib import Path
from unittest.mock import patch

import pandas as pd
from openpyxl import Workbook, load_workbook

import docatlas


def make_workbook_bytes(text: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Title"
    ws["A2"] = "DocAtlas"
    ws["B2"] = text
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def write_xlsx(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(make_workbook_bytes(text))


def read_jsonl_gz(path: Path) -> list[dict]:
    with gzip.open(path, "rt", encoding="utf-8") as fh:
        return [json.loads(line) for line in fh if line.strip()]


def dummy_cfg() -> docatlas.AzureConfig:
    return docatlas.AzureConfig(
        api_key="",
        chat_api_key="",
        embeddings_api_key="",
        api_version="",
        api_key_header="api-key",
        chat_base_url="",
        chat_path="",
        chat_deployment="",
        embeddings_base_url="",
        embeddings_path="",
        embeddings_deployment="",
        include_model_in_body=False,
    )


def sample_doc(
    doc_id: str,
    file_key: str,
    file_path: str,
    file_name: str | None = None,
    *,
    category: str = "Other",
    extraction_status: str = "ok",
    review_flags: str = "",
) -> docatlas.DocRecord:
    return docatlas.DocRecord(
        doc_id=doc_id,
        file_key=file_key,
        file_name=file_name or Path(file_path).name,
        file_path=file_path,
        source_path=file_path,
        file_ext=Path(file_path).suffix.lower() or ".xlsx",
        category=category,
        tags=["tag1", "tag2"],
        short_summary="Test summary",
        normalized_title="Test Product A12345 Technical Overview",
        long_summary="Longer test summary for workbook output.",
        word_count=25,
        char_count=180,
        extraction_status=extraction_status,
        review_flags=review_flags,
        duplicate_of="",
        duplicate_score=None,
        duplicate_group_id="",
        near_duplicate_of="",
        near_duplicate_score=None,
        near_duplicate_group_id="",
        review_group_id="",
        duplicate_relation_type="",
        moved_to="",
    )


def sample_article(doc_id: str, file_key: str, file_path: str) -> docatlas.ArticleRecord:
    return docatlas.ArticleRecord(
        doc_id=doc_id,
        file_key=file_key,
        file_name=Path(file_path).name,
        file_path=file_path,
        article_index=1,
        article_title="Article title",
        article_summary="Article summary",
        duplicate_of="",
        duplicate_score=None,
        duplicate_group_id="",
    )


class DocAtlasRegressionTests(unittest.TestCase):
    def make_tempdir(self) -> Path:
        temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(temp_dir.cleanup)
        return Path(temp_dir.name)

    def category_path_map(self) -> dict:
        return {
            "TestApp": {
                "Other": "/Life_Sciences/Life_Science_Applications/TestApp/Other",
                "Nanodrop": "/Life_Sciences/Life_Science_Applications/TestApp/Nanodrop",
                "Protein expression": "/Life_Sciences/Life_Science_Applications/TestApp/Protein_Expression",
                "Water": "/Life_Sciences/Life_Science_Applications/TestApp/Water",
                "Custom DNA Oligos": "/Life_Sciences/Life_Science_Applications/TestApp/Custom_DNA_Oligos",
                "Dynabeads": "/Life_Sciences/Life_Science_Applications/TestApp/Dynabeads",
            }
        }

    def prepare_logging(self, out_dir: Path) -> None:
        docatlas.setup_logging(out_dir)
        self.addCleanup(logging.shutdown)

    def test_write_unsupported_cleanup_workbook_formats_and_sorts(self) -> None:
        output_dir = self.make_tempdir()
        items = [
            docatlas.UnsupportedFileRecord("b.msg", "team/inbox/b.msg", ".msg", "file"),
            docatlas.UnsupportedFileRecord("a.url", "team/inbox/a.url", ".url", "file"),
            docatlas.UnsupportedFileRecord("inside.bin", "bundle.zip!/nested/inside.bin", ".bin", "zip_member"),
            docatlas.UnsupportedFileRecord("broken.zip", "broken.zip", ".zip", "invalid_zip"),
            docatlas.UnsupportedFileRecord("c.url", "other/c.url", ".url", "file"),
        ]

        workbook_path = docatlas.write_unsupported_cleanup_workbook(output_dir, items)

        self.assertTrue(workbook_path.exists())
        wb = load_workbook(workbook_path)
        self.assertEqual(wb.sheetnames, ["cleanup_queue", "cleanup_legend"])

        queue_ws = wb["cleanup_queue"]
        headers = [queue_ws.cell(row=1, column=c).value for c in range(1, queue_ws.max_column + 1)]
        self.assertEqual(
            headers,
            [
                "FileType",
                "FileName",
                "FilePath",
                "SourceKind",
                "SourceFolder",
                "RecommendedAction",
                "DeleteCandidate",
                "Decision",
                "DecisionNotes",
                "ReviewedBy",
                "FinalDisposition",
            ],
        )
        self.assertEqual(queue_ws.freeze_panes, "A2")
        self.assertEqual(queue_ws.auto_filter.ref, "A1:K6")
        self.assertEqual(len(queue_ws.data_validations.dataValidation), 1)
        self.assertTrue(len(queue_ws.conditional_formatting) >= 1)
        self.assertEqual(queue_ws["E2"].value, "team/inbox")
        self.assertEqual(queue_ws["E3"].value, "team/inbox")
        self.assertEqual(queue_ws["F2"].value, "Review")
        self.assertFalse(bool(queue_ws["G2"].value))
        legend_ws = wb["cleanup_legend"]
        self.assertEqual(legend_ws["A2"].value, "Keep")
        self.assertEqual(legend_ws["A5"].value, "Needs Follow-up")

    def test_process_file_supports_xls_via_conversion(self) -> None:
        root = self.make_tempdir()
        xls_path = root / "legacy.xls"
        xls_path.write_bytes(b"fake xls placeholder")
        converted_path = root / "converted.xlsx"
        write_xlsx(converted_path, " ".join(["legacy xls content"] * 20))

        with patch("docatlas.convert_xls_to_xlsx", return_value=converted_path):
            text, articles, status = docatlas.process_file(
                xls_path,
                ocrmypdf_enabled=False,
                articles_enabled=False,
                source_label="legacy.xls",
            )

        self.assertIn("legacy xls content", text)
        self.assertEqual(articles, [])
        self.assertEqual(status, "ok")

    def test_detect_duplicates_does_not_use_embedding_similarity_for_exact(self) -> None:
        duplicate_of, duplicate_score, duplicate_group = docatlas.detect_duplicates(
            [
                ("DOC-1", "hash-a", docatlas.np.array([1.0, 0.0], dtype=docatlas.np.float32)),
                ("DOC-2", "hash-b", docatlas.np.array([0.999, 0.001], dtype=docatlas.np.float32)),
            ],
            docatlas.DUPLICATE_THRESHOLD,
        )
        self.assertEqual(duplicate_of, {})
        self.assertEqual(duplicate_score, {})
        self.assertEqual(duplicate_group, {})

    def test_run_pipeline_failed_spreadsheets_do_not_collapse_into_exact_duplicates(self) -> None:
        root = self.make_tempdir()
        input_dir = root / "input"
        output_dir = root / "output"
        input_dir.mkdir(parents=True, exist_ok=True)
        (input_dir / "broken_a.xlsx").write_bytes(b"not an xlsx a")
        (input_dir / "broken_b.xlsx").write_bytes(b"not an xlsx b")

        docatlas.run_pipeline(
            input_dir=input_dir,
            output_dir=output_dir,
            categories=["Other"],
            cfg=dummy_cfg(),
            dry_run=True,
            use_resume=False,
            ocrmypdf_enabled=False,
            app_name="TestApp",
            embeddings_source=docatlas.EMBEDDINGS_SOURCE_NONE,
            append_excel=False,
            category_path_map=self.category_path_map(),
            include_full_text_output=False,
            no_move=True,
            articles_enabled=False,
        )
        self.addCleanup(logging.shutdown)

        docs_df = pd.read_excel(output_dir / "TestApp__docatlas_summaries.xlsx", sheet_name="Documents")
        self.assertEqual(len(docs_df), 2)
        self.assertTrue((docs_df["ExtractionStatus"] == "no_text").all())
        self.assertTrue((docs_df["DuplicateOf"].fillna("") == "").all())
        self.assertTrue((docs_df["DupScore"].fillna(0.0) == 0.0).all())

    def test_weak_near_duplicate_edges_require_structural_signal(self) -> None:
        docs = [
            sample_doc("DOC-1", "k1", "nanodrop/enablement_final.pptx", category="CatA"),
            sample_doc("DOC-2", "k2", "xenon/introduction_overview.pptx", category="CatA"),
            sample_doc("DOC-3", "k3", "crispr/product_alpha_overview.pdf", category="CatB"),
            sample_doc("DOC-4", "k4", "crispr/product_alpha_update.docx", category="CatB"),
        ]
        score = 0.80
        vec_a = docatlas.np.array([1.0, 0.0], dtype=docatlas.np.float32)
        vec_b = docatlas.np.array([score, (1 - score**2) ** 0.5], dtype=docatlas.np.float32)
        doc_embeddings = {
            "DOC-1": vec_a,
            "DOC-2": vec_b,
            "DOC-3": vec_a,
            "DOC-4": vec_b,
        }

        near_of, _near_score, _near_group, _near_adj, near_edges = docatlas.detect_near_duplicates_docs(docs, doc_embeddings)

        self.assertNotIn(("DOC-1", "DOC-2"), near_edges)
        self.assertIn(("DOC-3", "DOC-4"), near_edges)

    def test_quick_estimate_runtime_uses_latest_sibling_last_run_stats(self) -> None:
        root = self.make_tempdir()
        charter_dir = root / "output" / "protein_biology" / "charter"
        old_run = charter_dir / "2026-03-27_run01"
        new_run = charter_dir / "2026-03-30_run01"
        old_run.mkdir(parents=True, exist_ok=True)
        new_run.mkdir(parents=True, exist_ok=True)

        docatlas.save_last_run_stats(
            old_run,
            {
                "elapsed_sec": 1000.0,
                "processed_files": 100,
                "total_size_mb": 50.0,
                "ocr_enabled": True,
                "embeddings_source": "full_text",
                "chat_deployment": "gpt-5.2",
            },
        )

        est_sec, source, settings_match = docatlas.quick_estimate_runtime(
            {"count": 200, "total_size_mb": 120.0},
            new_run,
            ocrmypdf_enabled=True,
            embeddings_source="full_text",
            chat_deployment="gpt-5.2",
        )

        self.assertEqual(source, "baseline")
        self.assertTrue(settings_match)
        self.assertEqual(est_sec, 2400.0)

    def test_article_type_prefers_manuals_over_troubleshooting_content(self) -> None:
        label = docatlas.classify_article_type_by_content(
            "superscript_user_guide.pdf",
            "SuperScript User Guide",
            "Guide for setup and operation.",
            "Includes a troubleshooting section and error examples.",
            "This user guide explains setup, operation, troubleshooting, and maintenance.",
        )
        self.assertEqual(label, "Manuals and Guides")

        label = docatlas.classify_article_type_by_content(
            "qpcr_error_resolution.pdf",
            "qPCR Error Resolution",
            "How to resolve common issues.",
            "Focuses on failure causes and corrective actions.",
            "Troubleshooting errors, failures, and issue resolution steps.",
        )
        self.assertEqual(label, "Troubleshooting")

    def test_write_excels_excludes_unreadable_docs_from_import(self) -> None:
        out_dir = self.make_tempdir()
        self.prepare_logging(out_dir)
        good_doc = sample_doc("20260305111752-DOC-00001", "key-1", "sub/a.xlsx", category="Other")
        unreadable_doc = sample_doc(
            "20260305111752-DOC-00002",
            "key-2",
            "sub/b.xlsx",
            category=docatlas.UNREADABLE_CATEGORY,
            extraction_status="no_text",
            review_flags="low_text,short_text",
        )

        docatlas.write_excels(
            out_dir=out_dir,
            docs=[good_doc, unreadable_doc],
            articles=[],
            full_text_rows=[
                {"doc_id": good_doc.doc_id, "full_text": "good body text"},
                {"doc_id": unreadable_doc.doc_id, "full_text": "bad body text"},
            ],
            app_name="TestApp",
            append_excel=False,
            category_path_map=self.category_path_map(),
            include_full_text_output=False,
            articles_enabled=False,
        )

        import_df = pd.read_excel(out_dir / "TestApp__docatlas_import.xlsx", sheet_name="import")
        self.assertEqual(len(import_df), 1)
        self.assertEqual(str(import_df.loc[0, "Title"]), good_doc.normalized_title)
        self.assertEqual(str(import_df.loc[0, "Attachments"]), "./a.xlsx")

    def test_attachment_path_for_doc_uses_actual_filename(self) -> None:
        self.assertEqual(
            docatlas.attachment_path_for_doc("actual_file.pdf", "nested/ignored_name.docx"),
            "./actual_file.pdf",
        )
        self.assertEqual(
            docatlas.attachment_path_for_doc("", "folder/sub/backup.xlsx"),
            "./backup.xlsx",
        )

    def test_large_doc_summary_guard_uses_local_fallback(self) -> None:
        huge_text = ("important assay content " * 20000).strip()
        with patch("docatlas.summarize_document", side_effect=AssertionError("chat path should not run")):
            summary, flag = docatlas.summarize_document_safe(
                dummy_cfg(),
                huge_text,
                ["Other"],
                "huge.xlsx",
                "root/huge.xlsx",
            )

        self.assertEqual(flag, "summary_truncated_large_doc")
        self.assertTrue(summary["long_summary"])
        self.assertEqual(summary["category"], "Other")
        self.assertTrue(summary["normalized_title"])

    def test_normalize_import_title_text_strips_punctuation_and_questions(self) -> None:
        title = docatlas.normalize_import_title_text(
            "How does Expi293™ (A12345) improve yields?",
            file_name="Expi293_A12345.pdf",
        )
        self.assertEqual(title, "Expi293 A12345 Improve Yields")
        plus_title = docatlas.normalize_import_title_text(
            "EfficientFeed A+ B+ C+ stability guidance",
            file_name="efficientfeed.pdf",
        )
        self.assertEqual(plus_title, "EfficientFeed A Plus B Plus C Plus Stability Guidance")

    def test_summarize_with_model_normalizes_generated_title(self) -> None:
        payload = json.dumps(
            {
                "long_summary": "Long summary.",
                "short_summary": "Short summary.",
                "normalized_title": "Can Expi293™ A12345 improve yields?",
                "category": "Other",
                "tags": ["yield"],
            }
        )
        with patch("docatlas.call_azure_chat", return_value=payload):
            summary = docatlas.summarize_with_model(
                dummy_cfg(),
                "document text",
                ["Other"],
                file_name="Expi293_A12345.pdf",
                file_path="Cells/Expi293_A12345.pdf",
            )
        self.assertEqual(summary["normalized_title"], "Expi293 A12345 Improve Yields")

    def test_infer_category_uses_path_hints_for_nanodrop(self) -> None:
        category = docatlas._infer_category_from_text(
            "Training overview for a microvolume UV-Vis spectrophotometer and dealer enablement.",
            ["Nanodrop", "Other"],
            file_name="NanoDrop8_ProductAwareness_21SEP2021.pptx",
            file_path="Nanodrop/CONFIDENTIAL/NanoDrop Eight/NanoDrop8_ProductAwareness_21SEP2021.pptx",
        )
        self.assertEqual(category, "Nanodrop")

    def test_infer_category_uses_path_hints_for_protein_expression(self) -> None:
        category = docatlas._infer_category_from_text(
            "Training quiz covering promoters, cloning methods, and algal engineering.",
            ["Protein expression", "Other"],
            file_name="Algal Engineering_Prod_Quiz.docx",
            file_path="Protein expression/CONFIDENTIAL/Algae/Algal Engineering_Prod_Quiz.docx",
        )
        self.assertEqual(category, "Protein expression")

    def test_infer_category_prefers_custom_dna_oligos_over_water_for_oligo_coa(self) -> None:
        category = docatlas._infer_category_from_text(
            (
                "Certificate of analysis for custom DNA primers with sequences, "
                "well positions, extinction coefficient, GC content, and reconstitution guidance."
            ),
            ["Custom DNA Oligos", "Water", "Other"],
            file_name="18789122_e_plate.xls",
            file_path="PCR/Primers/Oligo Files/Randall-Primers/CoAs/18789122_e_plate.xls",
        )
        self.assertEqual(category, "Custom DNA Oligos")

    def test_infer_category_requires_explicit_water_signal(self) -> None:
        category = docatlas._infer_category_from_text(
            (
                "MSDS for Dynabeads anti-E.coli suspension with sodium phosphate buffer, "
                "safe handling guidance, and references to water samples."
            ),
            ["Dynabeads", "Water", "Other"],
            file_name="71003_MTR-NAIV_EN.pdf",
            file_path="Dynabeads/Dynabeads/CONFIDENTIAL/Bacterial Microbilogy/Documentation/MSDSs/71003_MTR-NAIV_EN.pdf",
        )
        self.assertEqual(category, "Dynabeads")

    def test_molecular_biology_config_includes_nanodrop_and_protein_expression(self) -> None:
        with open("applications.json", encoding="utf-8") as fh:
            app_config = json.load(fh)["applications"]
        with open("category_path_map.json", encoding="utf-8") as fh:
            path_map = json.load(fh)

        for app_name in ["Molecular Biology", "molbio"]:
            self.assertIn("Nanodrop", app_config[app_name])
            self.assertIn("Protein expression", app_config[app_name])
            self.assertIn("Nanodrop", path_map[app_name])
            self.assertIn("Protein expression", path_map[app_name])

    def test_cell_culture_config_includes_gibco_docs_and_cell_isolation(self) -> None:
        with open("applications.json", encoding="utf-8") as fh:
            app_config = json.load(fh)["applications"]
        with open("category_path_map.json", encoding="utf-8") as fh:
            path_map = json.load(fh)

        self.assertIn("Cell Isolation", app_config["Cell Culture"])
        self.assertIn("Gibco manufacturing and packaging docs", app_config["Cell Culture"])
        self.assertIn("Cell Isolation", path_map["Cell Culture"])
        self.assertIn("Gibco manufacturing and packaging docs", path_map["Cell Culture"])

    def test_cell_analysis_config_includes_hcs_and_tali(self) -> None:
        with open("applications.json", encoding="utf-8") as fh:
            app_config = json.load(fh)["applications"]
        with open("category_path_map.json", encoding="utf-8") as fh:
            path_map = json.load(fh)

        self.assertIn("HCS", app_config["Cell Analysis"])
        self.assertIn("Tali", app_config["Cell Analysis"])
        self.assertIn("HCS", path_map["Cell Analysis"])
        self.assertIn("Tali", path_map["Cell Analysis"])

    def test_protein_biology_config_includes_folder_priority_categories(self) -> None:
        with open("applications.json", encoding="utf-8") as fh:
            app_config = json.load(fh)["applications"]
        with open("category_path_map.json", encoding="utf-8") as fh:
            path_map = json.load(fh)

        self.assertIn("Protein Affinity Purification", app_config["Protein Biology"])
        self.assertIn("Protein Assays", app_config["Protein Biology"])
        self.assertIn("Western Blotting", app_config["Protein Biology"])
        self.assertIn("Protein Affinity Purification", path_map["Protein Biology"])
        self.assertIn("Protein Assays", path_map["Protein Biology"])
        self.assertIn("Western Blotting", path_map["Protein Biology"])

    def test_flowplex_config_includes_folder_priority_categories(self) -> None:
        with open("applications.json", encoding="utf-8") as fh:
            app_config = json.load(fh)["applications"]
        with open("category_path_map.json", encoding="utf-8") as fh:
            path_map = json.load(fh)

        self.assertIn("Flow Cytometry General", app_config["FlowPlex"])
        self.assertIn("Flow Cytometry Fluorochromes and Spectra", app_config["FlowPlex"])
        self.assertIn("Attune Xenith", app_config["FlowPlex"])
        self.assertIn("ProQuantum Protein Assays", app_config["FlowPlex"])
        self.assertIn("Flow Cytometry General", path_map["FlowPlex"])
        self.assertIn("Flow Cytometry Fluorochromes and Spectra", path_map["FlowPlex"])
        self.assertIn("Attune Xenith", path_map["FlowPlex"])
        self.assertIn("ProQuantum Protein Assays", path_map["FlowPlex"])

    def test_infer_category_uses_path_hints_for_gibco_docs(self) -> None:
        category = docatlas._infer_category_from_text(
            "Packaging specification and manufacturing label change guidance for Gibco product lines.",
            ["Gibco manufacturing and packaging docs", "Liquid Cell Culture", "Other"],
            file_name="Gibco_Label_Claim_Change.pdf",
            file_path="Gibco manufacturing and packaging docs/Gibco_Label_Claim_Change.pdf",
        )
        self.assertEqual(category, "Gibco manufacturing and packaging docs")

    def test_infer_category_prefers_cell_isolation_for_primary_cell_kit_content(self) -> None:
        category = docatlas._infer_category_from_text(
            "Primary cell isolation kit overview for neuron and cardiomyocyte isolation workflows.",
            ["Cell Isolation", "Dissociation Reagents", "Other"],
            file_name="Primary Cell Isolation Kits.ppt",
            file_path="Cell culture reagents antibiotics and supplements/Cell Isolation and Dissociation reagents/Primary Cell Isolation Kits.ppt",
        )
        self.assertEqual(category, "Cell Isolation")

    def test_infer_category_maps_extracellular_matrices_to_ecm_and_3d_culture(self) -> None:
        category = docatlas._infer_category_from_text(
            "Extracellular matrix hydrogel support for 3D culture and organoid workflows.",
            ["ECM and 3D Culture", "Other"],
            file_name="Basement_Membrane_Matrix_Guide.pdf",
            file_path="Extracellular matrices and 3D cultures/Basement_Membrane_Matrix_Guide.pdf",
        )
        self.assertEqual(category, "ECM and 3D Culture")

    def test_infer_category_maps_software_celleste_folder_to_celleste_category(self) -> None:
        category = docatlas._infer_category_from_text(
            "Release notes and image analysis software training materials for EVOS imaging workflows.",
            ["Celleste Image Analysis Software", "Other"],
            file_name="Celleste_Release_Notes.pdf",
            file_path="CELL ANALYSIS INSTRUMENTS and SOFTWARE/SOFTWARE - CELLESTE/CA EVOS Imaging Software Celleste - Release Notes/Celleste_Release_Notes.pdf",
        )
        self.assertEqual(category, "Celleste Image Analysis Software")

    def test_infer_category_maps_flowplex_bigfoot_typo_folder(self) -> None:
        category = docatlas._infer_category_from_text(
            "Application notes and sorter workflows for the Bigfoot cell sorter.",
            ["Bigfoot Applications", "Bigfoot Instruments and Accessories", "Other"],
            file_name="Bigfoot workflow note.pdf",
            file_path="FlowPlex/Flow Cytometry and Cell Sorting- CONFIDENTIAL/Cell Sorters/Bigfoot Applicationns/Bigfoot workflow note.pdf",
        )
        self.assertEqual(category, "Bigfoot Applications")

    def test_infer_category_maps_flowplex_attune_xenith_folder(self) -> None:
        category = docatlas._infer_category_from_text(
            "Setup and application materials for the Attune Xenith system.",
            ["Attune Xenith", "Attune CytPix", "Other"],
            file_name="Attune Xenith setup.pdf",
            file_path="FlowPlex/Flow Cytometry and Cell Sorting- CONFIDENTIAL/Flow Cytometers/Attune Instruments and Applications/Attune Xenith/Attune Xenith setup.pdf",
        )
        self.assertEqual(category, "Attune Xenith")

    def test_infer_category_maps_flowplex_fluorochrome_folder(self) -> None:
        category = docatlas._infer_category_from_text(
            "Emission and excitation spectra for fluorochromes used in flow cytometry assays.",
            ["Flow Cytometry Fluorochromes and Spectra", "Spectral Flow", "Other"],
            file_name="eFluor spectra.pdf",
            file_path="FlowPlex/Flow Cytometry and Cell Sorting- CONFIDENTIAL/Flow Cytometry Assays and Reagents/Flow Cytometry Fluorochromes and Spectra/eFluor spectra.pdf",
        )
        self.assertEqual(category, "Flow Cytometry Fluorochromes and Spectra")

    def test_infer_category_maps_hcs_folder_to_hcs(self) -> None:
        category = docatlas._infer_category_from_text(
            "Application note for high content screening assay setup and analysis workflows.",
            ["HCS", "Other"],
            file_name="HCS_Application_Note.pdf",
            file_path="CELL ANALYSIS INSTRUMENTS and SOFTWARE/INSTRUMENT - HCS/HCS - Application and Product Notes/HCS_Application_Note.pdf",
        )
        self.assertEqual(category, "HCS")

    def test_infer_category_maps_tali_folder_to_tali(self) -> None:
        category = docatlas._infer_category_from_text(
            "User guide for the image-based cytometer with setup, counting, and viability workflows.",
            ["Tali", "Other"],
            file_name="Tali_User_Guide.pdf",
            file_path="CELL ANALYSIS INSTRUMENTS and SOFTWARE/INSTRUMENT - TALI/Tali User Guides/Tali_User_Guide.pdf",
        )
        self.assertEqual(category, "Tali")

    def test_infer_category_maps_protein_alias_folders(self) -> None:
        category = docatlas._infer_category_from_text(
            "Troubleshooting and imaging guide for chemiluminescent blot detection.",
            ["iBright and my ECL", "Western Blotting", "Other"],
            file_name="iBright_and_myECL_Troubleshooting.pdf",
            file_path="iBright and myECL/Troubleshooting/iBright_and_myECL_Troubleshooting.pdf",
        )
        self.assertEqual(category, "iBright and my ECL")

    def test_infer_category_maps_protein_broad_parent_folders(self) -> None:
        category = docatlas._infer_category_from_text(
            "Application note for protein A and protein G agarose affinity purification workflows.",
            ["Protein Affinity Purification", "POROS and CaptureSelect", "Resins", "Other"],
            file_name="Protein_G_Agarose_Workflow.pdf",
            file_path="Protein Affinity Purification/Application and Product Notes/Protein_G_Agarose_Workflow.pdf",
        )
        self.assertEqual(category, "Protein Affinity Purification")

    def test_infer_category_maps_protein_assays_folder(self) -> None:
        category = docatlas._infer_category_from_text(
            "Rapid Gold BCA protocol for NanoDrop protein quantification.",
            ["Protein Assays", "Colorimetric Protein Assays", "Fluorescent Protein Assays", "Other"],
            file_name="nanodrop-one-onec-rapid-gold-bca-protocol-en.pdf",
            file_path="Protein Assays/Application and Product Notes/nanodrop-one-onec-rapid-gold-bca-protocol-en.pdf",
        )
        self.assertEqual(category, "Protein Assays")

    def test_infer_category_maps_western_blotting_folder(self) -> None:
        category = docatlas._infer_category_from_text(
            "Western blot troubleshooting guide for membrane blocking and detection.",
            ["Western Blotting", "Other"],
            file_name="Western_Blot_Troubleshooting.pdf",
            file_path="Western Blotting/Troubleshooting/Western_Blot_Troubleshooting.pdf",
        )
        self.assertEqual(category, "Western Blotting")

    def test_list_files_discovers_zip_members_with_logical_paths(self) -> None:
        input_dir = self.make_tempdir()
        write_xlsx(
            input_dir / "sub" / "direct.xlsx",
            " ".join(["direct workbook content"] * 20),
        )
        (input_dir / "sub" / "skip.url").write_text("[InternetShortcut]\nURL=https://example.com\n", encoding="utf-8")
        bundle_path = input_dir / "bundle.zip"
        with zipfile.ZipFile(bundle_path, "w") as zf:
            zf.writestr(
                "nested/inside.xlsx",
                make_workbook_bytes(" ".join(["zipped workbook content"] * 20)),
            )
            zf.writestr("nested/clip.mp4", b"not really a movie")

        files, unsupported, staged = docatlas.list_files(input_dir)
        try:
            display_paths = [item.display_path for item in files]
            self.assertEqual(
                display_paths,
                ["bundle.zip!/nested/inside.xlsx", "sub/direct.xlsx"],
            )
            unsupported_map = {(item.file_type, item.file_path, item.source_kind) for item in unsupported}
            self.assertIn((".url", "sub/skip.url", "file"), unsupported_map)
            self.assertIn((".mp4", "bundle.zip!/nested/clip.mp4", "zip_member"), unsupported_map)
        finally:
            if staged is not None:
                staged.cleanup()

    def test_list_files_records_invalid_zip_as_unsupported(self) -> None:
        input_dir = self.make_tempdir()
        (input_dir / "broken.zip").write_text("not a real zip", encoding="utf-8")

        files, unsupported, staged = docatlas.list_files(input_dir)
        try:
            self.assertEqual(files, [])
            self.assertEqual(len(unsupported), 1)
            self.assertEqual(unsupported[0].file_path, "broken.zip")
            self.assertEqual(unsupported[0].source_kind, "invalid_zip")
        finally:
            if staged is not None:
                staged.cleanup()

    def test_build_import_path_uses_article_type_slug(self) -> None:
        path = docatlas.build_import_path(
            "/Life_Sciences/Life_Science_Applications/Quantitative_PCR/Real-Time_PCR/TaqMan_Assays_and_Applications/TaqMan_Protein_Assay",
            "TaqMan Protein Assay",
            "Application and Product Notes",
        )
        self.assertEqual(
            path,
            "/Life_Sciences/Life_Science_Applications/Quantitative_PCR/Real-Time_PCR/TaqMan_Assays_and_Applications/TaqMan_Protein_Assay/TaqMan_Protein_Assay_Application_and_Product_Notes",
        )

    def test_write_excels_dedupes_append_across_absolute_and_relative_paths(self) -> None:
        out_dir = self.make_tempdir()
        self.prepare_logging(out_dir)
        peers_path = out_dir / "TestApp__docatlas_summaries.xlsx"
        with pd.ExcelWriter(peers_path, engine="openpyxl") as writer:
            pd.DataFrame(
                [{"Category": "Other", "FilePath": "C:/root/input/sub/a.xlsx", "FileName": "a.xlsx"}]
            ).to_excel(writer, index=False, sheet_name="Documents")
            pd.DataFrame(columns=["ReviewGroupID", "Category", "FilePath", "FileName"]).to_excel(
                writer,
                index=False,
                sheet_name="Duplicates",
            )

        doc = sample_doc("20260305111752-DOC-00001", "new-key", "sub/a.xlsx")
        docatlas.write_excels(
            out_dir=out_dir,
            docs=[doc],
            articles=[],
            full_text_rows=[{"doc_id": doc.doc_id, "full_text": "body text"}],
            app_name="TestApp",
            append_excel=True,
            category_path_map=self.category_path_map(),
            include_full_text_output=False,
            articles_enabled=False,
        )

        docs_df = pd.read_excel(peers_path, sheet_name="Documents")
        self.assertEqual(len(docs_df), 1)
        self.assertEqual(docs_df.loc[0, "FilePath"], "C:/root/input/sub/a.xlsx")
        archive_path = out_dir / "TestApp__docatlas_full_text.jsonl.gz"
        self.assertTrue(archive_path.exists())
        self.assertFalse((out_dir / "TestApp__docatlas_full_text.xlsx").exists())

    def test_write_excels_writes_full_text_archive_by_default_and_dedupes_append(self) -> None:
        out_dir = self.make_tempdir()
        self.prepare_logging(out_dir)
        first_doc = sample_doc("20260305111752-DOC-00001", "key-1", "sub/a.xlsx")
        second_doc = sample_doc("20260305111752-DOC-00002", "key-2", "sub/b.xlsx")

        docatlas.write_excels(
            out_dir=out_dir,
            docs=[first_doc],
            articles=[],
            full_text_rows=[
                {
                    "doc_id": first_doc.doc_id,
                    "file_key": first_doc.file_key,
                    "file_name": first_doc.file_name,
                    "file_path": first_doc.file_path,
                    "category": first_doc.category,
                    "short_summary": first_doc.short_summary,
                    "long_summary": first_doc.long_summary,
                    "tags": ", ".join(first_doc.tags),
                    "word_count": first_doc.word_count,
                    "char_count": first_doc.char_count,
                    "extraction_status": first_doc.extraction_status,
                    "review_flags": first_doc.review_flags,
                    "moved_to": first_doc.moved_to,
                    "full_text": "first body text",
                }
            ],
            app_name="TestApp",
            append_excel=False,
            category_path_map=self.category_path_map(),
            include_full_text_output=False,
            articles_enabled=False,
        )

        archive_path = out_dir / "TestApp__docatlas_full_text.jsonl.gz"
        records = read_jsonl_gz(archive_path)
        self.assertEqual(len(records), 1)
        self.assertEqual(records[0]["file_key"], "key-1")
        self.assertEqual(records[0]["full_text"], "first body text")

        docatlas.write_excels(
            out_dir=out_dir,
            docs=[first_doc, second_doc],
            articles=[],
            full_text_rows=[
                {
                    "doc_id": first_doc.doc_id,
                    "file_key": first_doc.file_key,
                    "file_name": first_doc.file_name,
                    "file_path": first_doc.file_path,
                    "category": first_doc.category,
                    "short_summary": first_doc.short_summary,
                    "long_summary": first_doc.long_summary,
                    "tags": ", ".join(first_doc.tags),
                    "word_count": first_doc.word_count,
                    "char_count": first_doc.char_count,
                    "extraction_status": first_doc.extraction_status,
                    "review_flags": first_doc.review_flags,
                    "moved_to": first_doc.moved_to,
                    "full_text": "first body text again",
                },
                {
                    "doc_id": second_doc.doc_id,
                    "file_key": second_doc.file_key,
                    "file_name": second_doc.file_name,
                    "file_path": second_doc.file_path,
                    "category": second_doc.category,
                    "short_summary": second_doc.short_summary,
                    "long_summary": second_doc.long_summary,
                    "tags": ", ".join(second_doc.tags),
                    "word_count": second_doc.word_count,
                    "char_count": second_doc.char_count,
                    "extraction_status": second_doc.extraction_status,
                    "review_flags": second_doc.review_flags,
                    "moved_to": second_doc.moved_to,
                    "full_text": "second body text",
                },
            ],
            app_name="TestApp",
            append_excel=True,
            category_path_map=self.category_path_map(),
            include_full_text_output=False,
            articles_enabled=False,
        )

        records = read_jsonl_gz(archive_path)
        self.assertEqual([record["file_key"] for record in records], ["key-1", "key-2"])
        self.assertEqual(records[1]["full_text"], "second body text")

    def test_write_excels_strips_duplicate_header_rows_on_append(self) -> None:
        out_dir = self.make_tempdir()
        self.prepare_logging(out_dir)
        doc = sample_doc("20260305111752-DOC-00001", "key-1", "sub/a.xlsx")
        docatlas.write_excels(
            out_dir=out_dir,
            docs=[doc],
            articles=[],
            full_text_rows=[{"doc_id": doc.doc_id, "file_key": doc.file_key, "file_name": doc.file_name, "file_path": doc.file_path, "category": doc.category, "short_summary": doc.short_summary, "long_summary": doc.long_summary, "tags": ", ".join(doc.tags), "word_count": doc.word_count, "char_count": doc.char_count, "extraction_status": doc.extraction_status, "review_flags": doc.review_flags, "moved_to": doc.moved_to, "full_text": "body text"}],
            app_name="TestApp",
            append_excel=False,
            category_path_map=self.category_path_map(),
            include_full_text_output=False,
            articles_enabled=False,
        )

        peers_path = out_dir / "TestApp__docatlas_summaries.xlsx"
        import_path = out_dir / "TestApp__docatlas_import.xlsx"

        wb = load_workbook(peers_path)
        ws = wb["Documents"]
        ws.append([cell.value for cell in ws[1]])
        wb.save(peers_path)
        wb.close()

        wb = load_workbook(import_path)
        ws = wb["import"]
        ws.append([cell.value for cell in ws[1]])
        wb.save(import_path)
        wb.close()

        docatlas.write_excels(
            out_dir=out_dir,
            docs=[],
            articles=[],
            full_text_rows=[],
            app_name="TestApp",
            append_excel=True,
            category_path_map=self.category_path_map(),
            include_full_text_output=False,
            articles_enabled=False,
        )

        wb = load_workbook(peers_path, read_only=True, data_only=True)
        ws = wb["Documents"]
        rows = list(ws.iter_rows(values_only=True))
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[1][1], "sub/a.xlsx")
        wb.close()

        wb = load_workbook(import_path, read_only=True, data_only=True)
        ws = wb["import"]
        rows = list(ws.iter_rows(values_only=True))
        self.assertEqual(len(rows), 2)
        self.assertEqual(
            rows[1][1],
            "/Life_Sciences/Life_Science_Applications/TestApp/Other/Other_Documentation",
        )
        wb.close()

    def test_write_excels_omits_articles_sheet_when_disabled(self) -> None:
        out_dir = self.make_tempdir()
        self.prepare_logging(out_dir)
        doc = sample_doc("20260305111752-DOC-00001", "key-1", "sub/a.xlsx")
        docatlas.write_excels(
            out_dir=out_dir,
            docs=[doc],
            articles=[sample_article(doc.doc_id, doc.file_key, doc.file_path)],
            full_text_rows=[{"doc_id": doc.doc_id, "full_text": "body text"}],
            app_name="TestApp",
            append_excel=False,
            category_path_map=self.category_path_map(),
            include_full_text_output=False,
            articles_enabled=False,
        )

        wb = load_workbook(out_dir / "TestApp__docatlas_summaries.xlsx")
        self.assertEqual(wb.sheetnames, ["Documents", "Duplicates", "duplicate_review_legend"])
        self.assertEqual(
            [cell.value for cell in wb["Duplicates"][1]][-4:],
            ["GroupReviewed", "Decision", "DecisionNotes", "ReviewedBy"],
        )
        self.assertEqual(list(wb["Documents"].tables.keys()), [])
        self.assertEqual(list(wb["Duplicates"].tables.keys()), [])
        self.assertTrue((out_dir / "TestApp__docatlas_full_text.jsonl.gz").exists())
        self.assertFalse((out_dir / "TestApp__docatlas_full_text.xlsx").exists())

    def test_write_excels_preserves_existing_articles_sheet_on_append(self) -> None:
        out_dir = self.make_tempdir()
        self.prepare_logging(out_dir)
        first_doc = sample_doc("20260305111752-DOC-00001", "key-1", "sub/a.xlsx")
        first_article = sample_article(first_doc.doc_id, first_doc.file_key, first_doc.file_path)
        docatlas.write_excels(
            out_dir=out_dir,
            docs=[first_doc],
            articles=[first_article],
            full_text_rows=[{"doc_id": first_doc.doc_id, "full_text": "body text"}],
            app_name="TestApp",
            append_excel=False,
            category_path_map=self.category_path_map(),
            include_full_text_output=False,
            articles_enabled=True,
        )

        second_doc = sample_doc("20260305111752-DOC-00002", "key-2", "sub/b.xlsx")
        docatlas.write_excels(
            out_dir=out_dir,
            docs=[second_doc],
            articles=[],
            full_text_rows=[{"doc_id": second_doc.doc_id, "full_text": "body text"}],
            app_name="TestApp",
            append_excel=True,
            category_path_map=self.category_path_map(),
            include_full_text_output=False,
            articles_enabled=False,
        )

        peers_path = out_dir / "TestApp__docatlas_summaries.xlsx"
        wb = load_workbook(peers_path)
        self.assertIn("Articles", wb.sheetnames)
        articles_df = pd.read_excel(peers_path, sheet_name="Articles")
        self.assertEqual(len(articles_df), 1)
        self.assertEqual(str(articles_df.loc[0, "ParentDocID"]), first_doc.doc_id)

    def test_run_pipeline_serial_uses_relative_and_zip_logical_paths(self) -> None:
        root = self.make_tempdir()
        input_dir = root / "input"
        output_dir = root / "output"
        write_xlsx(
            input_dir / "sub" / "direct.xlsx",
            " ".join(["serial pipeline content application note guide"] * 20),
        )
        with zipfile.ZipFile(input_dir / "bundle.zip", "w") as zf:
            zf.writestr(
                "nested/inside.xlsx",
                make_workbook_bytes(" ".join(["serial zipped content"] * 25)),
            )

        docatlas.run_pipeline(
            input_dir=input_dir,
            output_dir=output_dir,
            categories=["Other"],
            cfg=dummy_cfg(),
            dry_run=True,
            use_resume=False,
            ocrmypdf_enabled=False,
            app_name="TestApp",
            embeddings_source="document",
            append_excel=False,
            category_path_map=self.category_path_map(),
            include_full_text_output=False,
            no_move=True,
            articles_enabled=False,
        )
        self.addCleanup(logging.shutdown)

        peers_path = output_dir / "TestApp__docatlas_summaries.xlsx"
        wb = load_workbook(peers_path)
        self.assertEqual(wb.sheetnames, ["Documents", "Duplicates", "duplicate_review_legend"])
        self.assertTrue((output_dir / "TestApp__docatlas_full_text.jsonl.gz").exists())
        self.assertFalse((output_dir / "TestApp__docatlas_full_text.xlsx").exists())
        self.assertFalse((output_dir / "unsupported_cleanup.xlsx").exists())
        docs_df = pd.read_excel(peers_path, sheet_name="Documents")
        self.assertEqual(
            sorted(docs_df["FilePath"].astype(str).tolist()),
            ["bundle.zip!/nested/inside.xlsx", "sub/direct.xlsx"],
        )
        summary_text = (output_dir / "summary_report.txt").read_text(encoding="utf-8")
        unsupported_text = (output_dir / "unsupported_files_report.txt").read_text(encoding="utf-8")
        self.assertIn("total_unsupported_files: 0", summary_text)
        self.assertIn("Total unsupported files: 0", unsupported_text)

    def test_run_pipeline_parallel_writes_no_articles_sheet(self) -> None:
        root = self.make_tempdir()
        input_dir = root / "input"
        output_dir = root / "output"
        write_xlsx(
            input_dir / "a" / "one.xlsx",
            " ".join(["parallel pipeline content troubleshooting"] * 20),
        )
        with zipfile.ZipFile(input_dir / "bundle.zip", "w") as zf:
            zf.writestr(
                "nested/two.xlsx",
                make_workbook_bytes(" ".join(["parallel zipped content"] * 25)),
            )

        docatlas.run_pipeline_parallel(
            input_dir=input_dir,
            output_dir=output_dir,
            categories=["Other"],
            cfg=dummy_cfg(),
            dry_run=True,
            use_resume=False,
            ocrmypdf_enabled=False,
            app_name="TestApp",
            embeddings_source="document",
            append_excel=False,
            workers=2,
            category_path_map=self.category_path_map(),
            include_full_text_output=False,
            no_move=True,
            articles_enabled=False,
        )
        self.addCleanup(logging.shutdown)

        peers_path = output_dir / "TestApp__docatlas_summaries.xlsx"
        wb = load_workbook(peers_path)
        self.assertEqual(wb.sheetnames, ["Documents", "Duplicates", "duplicate_review_legend"])
        self.assertTrue((output_dir / "TestApp__docatlas_full_text.jsonl.gz").exists())
        self.assertFalse((output_dir / "TestApp__docatlas_full_text.xlsx").exists())
        docs_df = pd.read_excel(peers_path, sheet_name="Documents")
        self.assertEqual(
            sorted(docs_df["FilePath"].astype(str).tolist()),
            ["a/one.xlsx", "bundle.zip!/nested/two.xlsx"],
        )

    def test_write_excels_core_duplicate_review_workflow_is_present(self) -> None:
        out_dir = self.make_tempdir()
        self.prepare_logging(out_dir)
        docs = [
            sample_doc("20260305111754-DOC-00001", "key-1", "sub/a.xlsx", category="Other"),
            sample_doc("20260305111754-DOC-00002", "key-2", "sub/b.xlsx", category="Other"),
        ]
        docs[0].review_group_id = "RGRP-0001"
        docs[0].near_duplicate_of = "DOC-00002"
        docs[0].near_dup_score = 0.91
        docs[0].duplicate_relation_type = "near"
        docs[1].review_group_id = "RGRP-0001"
        docs[1].near_duplicate_of = "DOC-00001"
        docs[1].near_dup_score = 0.91
        docs[1].duplicate_relation_type = "near"

        docatlas.write_excels(
            out_dir=out_dir,
            docs=docs,
            articles=[],
            full_text_rows=[],
            app_name="TestApp",
            append_excel=False,
            category_path_map=self.category_path_map(),
            include_full_text_output=False,
            articles_enabled=False,
        )

        peers_path = out_dir / "TestApp__docatlas_summaries.xlsx"
        wb = load_workbook(peers_path)
        self.assertIn("duplicate_review_legend", wb.sheetnames)
        dup_ws = wb["Duplicates"]
        self.assertEqual(
            [cell.value for cell in dup_ws[1]][-4:],
            ["GroupReviewed", "Decision", "DecisionNotes", "ReviewedBy"],
        )
        self.assertEqual(dup_ws.freeze_panes, "A2")
        self.assertIsNotNone(dup_ws.auto_filter.ref)
        self.assertTrue(len(dup_ws.conditional_formatting) >= 1)
        self.assertEqual(list(wb["Documents"].tables.keys()), [])
        self.assertEqual(list(dup_ws.tables.keys()), [])
        legend_ws = wb["duplicate_review_legend"]
        self.assertEqual([cell.value for cell in legend_ws[1]], ["Field", "Allowed Values", "Meaning"])
        dv_formulas = [dv.formula1 for dv in dup_ws.data_validations.dataValidation]
        self.assertIn('"Reviewed,Unfinished"', dv_formulas)
        self.assertIn('"Primary,Keep,Drop,Needs Review"', dv_formulas)
        wb.close()

    def test_run_pipeline_writes_unsupported_reports_for_only_skipped_files(self) -> None:
        root = self.make_tempdir()
        input_dir = root / "input"
        output_dir = root / "output"
        input_dir.mkdir(parents=True, exist_ok=True)
        (input_dir / "link.url").write_text("[InternetShortcut]\nURL=https://example.com\n", encoding="utf-8")
        (input_dir / "movie.mp4").write_bytes(b"fake")
        (input_dir / "broken.zip").write_text("not a zip", encoding="utf-8")

        docatlas.run_pipeline(
            input_dir=input_dir,
            output_dir=output_dir,
            categories=["Other"],
            cfg=dummy_cfg(),
            dry_run=True,
            use_resume=False,
            ocrmypdf_enabled=False,
            app_name="TestApp",
            embeddings_source="document",
            append_excel=False,
            category_path_map=self.category_path_map(),
            include_full_text_output=False,
            no_move=True,
            articles_enabled=False,
        )
        self.addCleanup(logging.shutdown)

        self.assertFalse((output_dir / "TestApp__docatlas_summaries.xlsx").exists())
        summary_text = (output_dir / "summary_report.txt").read_text(encoding="utf-8")
        unsupported_text = (output_dir / "unsupported_files_report.txt").read_text(encoding="utf-8")
        cleanup_wb = load_workbook(output_dir / "unsupported_cleanup.xlsx")
        self.assertIn("total_unsupported_files: 3", summary_text)
        self.assertIn("- .mp4: 1", summary_text)
        self.assertIn("- .url: 1", summary_text)
        self.assertIn("- .zip: 1", summary_text)
        self.assertIn("Detailed List:", unsupported_text)
        self.assertIn(".url | link.url | link.url", unsupported_text)
        self.assertIn(".zip | broken.zip | broken.zip", unsupported_text)
        self.assertEqual(cleanup_wb.sheetnames, ["cleanup_queue", "cleanup_legend"])
        queue_df = pd.read_excel(output_dir / "unsupported_cleanup.xlsx", sheet_name="cleanup_queue")
        self.assertEqual(len(queue_df), 3)
        self.assertEqual(queue_df["RecommendedAction"].tolist(), ["Review", "Review", "Review"])
        self.assertTrue((queue_df["DeleteCandidate"] == False).all())

    def test_run_pipeline_writes_unsupported_cleanup_workbook_for_mixed_input(self) -> None:
        root = self.make_tempdir()
        input_dir = root / "input"
        output_dir = root / "output"
        write_xlsx(
            input_dir / "sub" / "direct.xlsx",
            " ".join(["mixed pipeline content application note guide"] * 20),
        )
        (input_dir / "sub" / "shortcut.url").write_text("[InternetShortcut]\nURL=https://example.com\n", encoding="utf-8")
        with zipfile.ZipFile(input_dir / "bundle.zip", "w") as zf:
            zf.writestr(
                "nested/inside.xlsx",
                make_workbook_bytes(" ".join(["mixed zipped content"] * 25)),
            )
            zf.writestr("nested/inside.url", "[InternetShortcut]\nURL=https://example.com\n")

        docatlas.run_pipeline(
            input_dir=input_dir,
            output_dir=output_dir,
            categories=["Other"],
            cfg=dummy_cfg(),
            dry_run=True,
            use_resume=False,
            ocrmypdf_enabled=False,
            app_name="TestApp",
            embeddings_source="document",
            append_excel=False,
            category_path_map=self.category_path_map(),
            include_full_text_output=False,
            no_move=True,
            articles_enabled=False,
        )
        self.addCleanup(logging.shutdown)

        self.assertTrue((output_dir / "TestApp__docatlas_summaries.xlsx").exists())
        cleanup_df = pd.read_excel(output_dir / "unsupported_cleanup.xlsx", sheet_name="cleanup_queue")
        self.assertEqual(len(cleanup_df), 2)
        self.assertEqual(set(cleanup_df["SourceKind"].astype(str)), {"file", "zip_member"})
        self.assertIn("sub", set(cleanup_df["SourceFolder"].astype(str)))
        self.assertIn("bundle.zip!/nested", set(cleanup_df["SourceFolder"].astype(str)))

    def test_resolve_cli_interactive_inputs_prompts_for_missing_values(self) -> None:
        root = self.make_tempdir()
        input_dir = root / "input"
        input_dir.mkdir(parents=True, exist_ok=True)
        output_dir = root / "output"

        args = Namespace(
            input=str(input_dir),
            output=None,
            categories=None,
            app="TestApp",
            charter_mode=False,
            signal_scan=False,
            no_move=False,
            no_ocrmypdf=False,
            embeddings_source=None,
            overwrite_excel=False,
            articles=False,
            workers=1,
        )

        with patch("sys.stdin.isatty", return_value=True):
            with patch(
                "builtins.input",
                side_effect=[
                    str(output_dir),
                    "",
                    "",
                    "",
                    "",
                    "y",
                    "2",
                ],
            ):
                result = docatlas.resolve_cli_interactive_inputs(args, [], {"TestApp": ["Other"]})

        (
            resolved_input,
            resolved_output,
            categories,
            app_name,
            ocrmypdf_enabled,
            embeddings_source,
            append_excel,
            no_move,
            articles_enabled,
            workers,
        ) = result
        self.assertEqual(resolved_input, input_dir)
        self.assertEqual(resolved_output, output_dir)
        self.assertEqual(categories, ["Other"])
        self.assertEqual(app_name, "TestApp")
        self.assertTrue(ocrmypdf_enabled)
        self.assertEqual(embeddings_source, "full_text")
        self.assertTrue(append_excel)
        self.assertTrue(no_move)
        self.assertTrue(articles_enabled)
        self.assertEqual(workers, 2)

    def test_build_app_folder_structure_script_creates_expected_layout(self) -> None:
        root = self.make_tempdir()
        base_dir = root / "DocAtlas"
        config_path = root / "applications.json"
        config_path.write_text(
            json.dumps({"applications": {"Molecular Biology": ["Other"], "qPCR": ["Other"]}}),
            encoding="utf-8",
        )

        result = subprocess.run(
            [
                "python",
                "build_app_folder_structure.py",
                "--base",
                str(base_dir),
                "--config",
                str(config_path),
            ],
            cwd=Path(__file__).resolve().parents[1],
            capture_output=True,
            text=True,
            check=True,
        )

        self.assertTrue((base_dir / "input" / "molecular_biology").exists())
        self.assertTrue((base_dir / "output" / "molecular_biology" / "charter").exists())
        self.assertTrue((base_dir / "output" / "molecular_biology" / "atlas").exists())
        self.assertTrue((base_dir / "input" / "qpcr").exists())
        self.assertTrue((base_dir / "archive" / "zips").exists())
        self.assertIn("Molecular Biology -> molecular_biology", result.stdout)
        self.assertIn("qPCR -> qpcr", result.stdout)


if __name__ == "__main__":
    unittest.main()
